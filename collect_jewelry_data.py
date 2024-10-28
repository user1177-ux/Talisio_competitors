from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import requests
from io import BytesIO
from openpyxl.worksheet.hyperlink import Hyperlink

# Функция для сбора данных с указанной категории
def collect_data_for_category(driver, wait, base_url, category_name):
    data = []
    
    print(f"Сбор данных для категории: {category_name}")
    driver.get(base_url)
    time.sleep(5)  # Небольшая задержка для загрузки контента
    
    while True:
        # Поиск всех элементов товаров на странице
        product_elements = driver.find_elements(By.XPATH, "//article[@class='card']")
        print(f"Найдено {len(product_elements)} товаров на странице.")

        # Ограничиваем обработку только первыми 5 товарами
        for index in range(min(5, len(product_elements))):
            try:
                print(f"Обрабатываю товар {index + 1}...")

                # Заново находим элементы, чтобы избежать ошибки stale element
                product_elements = driver.find_elements(By.XPATH, "//article[@class='card']")
                product = product_elements[index]

                # Получение ссылки на изображение
                try:
                    img_element = product.find_element(By.XPATH, ".//img")
                    img_url = img_element.get_attribute("src")
                    print(f"URL изображения: {img_url}")
                except Exception as e:
                    print(f"Ошибка при получении URL изображения: {str(e)}")
                    img_url = None

                # Нахождение и клик по названию для открытия товара
                try:
                    title_link = product.find_element(By.TAG_NAME, "a").get_attribute("href")
                    print(f"Открытие товара по ссылке: {title_link}")
                    
                    # Переход на страницу товара
                    driver.get(title_link)
                    time.sleep(5)

                    # Обновляем страницу после перехода
                    driver.refresh()
                    time.sleep(5)

                    # Получение названия товара (title)
                    try:
                        title_element = driver.find_element(By.XPATH, "//span[@class='title-h3 product-title']")
                        title = title_element.text
                        print(f"Название товара (title): {title}")
                    except Exception as e:
                        print(f"Ошибка при получении названия товара: {str(e)}")
                        title = "Название не найдено"

                    # Получение subtitle
                    try:
                        subtitle_element = driver.find_element(By.XPATH, "//h1[@class='product-subtitle']")
                        subtitle = subtitle_element.text
                        print(f"Subtitle товара: {subtitle}")
                    except Exception as e:
                        print(f"Ошибка при получении subtitle: {str(e)}")
                        subtitle = "Subtitle не найден"

                    # Получение цены товара
                    try:
                        price_element = driver.find_element(By.XPATH, "//p[@class='price']")
                        price = price_element.text
                        print(f"Цена товара: {price}")
                    except Exception as e:
                        print(f"Ошибка при получении цены товара: {str(e)}")
                        price = "Цена не найдена"

                    # Нажатие на кнопку "Item details" для раскрытия информации о деталях
                    try:
                        details_button = driver.find_element(By.XPATH, "//h2[contains(text(), 'Item details')]/..")
                        driver.execute_script("arguments[0].click();", details_button)  # Выполняем клик через JavaScript
                        time.sleep(2)  # Небольшая задержка, чтобы информация загрузилась
                        print("Кнопка для раскрытия деталей нажата.")
                    except Exception as e:
                        print(f"Ошибка при нажатии на кнопку для раскрытия деталей: {str(e)}")

                    # Получение деталей товара из блока с классом "content" и тегом <p>
                    try:
                        details_element = driver.find_element(By.XPATH, "//div[@class='content']//p")
                        details = details_element.text
                        print(f"Детали товара: {details}")
                    except Exception as e:
                        print(f"Ошибка при получении деталей товара: {str(e)}")
                        details = "Детали не найдены"

                    # Проверка корректности данных перед сохранением
                    if title and subtitle and price and details:
                        data.append({
                            "category": category_name,
                            "url": title_link,
                            "title": title,
                            "subtitle": subtitle,
                            "price": price,
                            "details": details,
                            "photo": img_url  # Сохраняем URL изображения
                        })
                    else:
                        print(f"Данные для товара {index + 1} не полные, пропускаем запись")

                except Exception as e:
                    print(f"Ошибка при клике по названию товара {index + 1}: {str(e)}")

                # Возврат на предыдущую страницу
                driver.back()
                time.sleep(5)

            except Exception as e:
                print(f"Ошибка при обработке товара {index + 1}: {str(e)}")

        # Проверка на наличие кнопки следующей страницы
        try:
            next_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@class='next']")))
            next_button.click()
            time.sleep(5)
        except:
            # Если кнопки нет, выходим из цикла
            print(f"Нет больше страниц для категории {category_name}")
            break
    
    return data

# Функция для вставки изображения в ячейку и изменение размера
def add_image_to_excel(ws, img_url, cell):
    try:
        response = requests.get(img_url)
        img_data = BytesIO(response.content)
        img = Image(img_data)

        # Увеличиваем размеры ячейки под изображение
        ws.row_dimensions[cell.row].height = 150  # Высота ячейки
        ws.column_dimensions[cell.column_letter].width = 25  # Ширина ячейки

        # Увеличиваем размеры изображения
        img.width = 150
        img.height = 150

        ws.add_image(img, cell.coordinate)
    except Exception as e:
        print(f"Ошибка при вставке изображения: {str(e)}")

# Основная функция для вставки данных и форматирования
def save_data_to_excel(all_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jewelry Data"
    
    # Заголовки
    ws.append(["Category", "URL", "Photo", "Title", "Subtitle", "Price", "Details"])
    
    # Вставка данных
    for row, item in enumerate(all_data, start=2):
        # Добавляем все данные в одну строку
        ws[f"A{row}"] = item["category"]
        ws[f"B{row}"] = item["url"]
        ws[f"D{row}"] = item["title"]
        ws[f"E{row}"] = item["subtitle"]
        ws[f"F{row}"] = item["price"]
        ws[f"G{row}"] = item["details"]

        # Добавление гиперссылки в URL
        url_cell = ws[f"B{row}"]
        ws[f"B{row}"].hyperlink = item["url"]

        # Вставляем изображение
        if item["photo"]:
            add_image_to_excel(ws, item["photo"], ws[f"C{row}"])

        # Центрируем остальные ячейки по вертикали и горизонтали
        for col in range(1, 7):
            cell = ws[f"{get_column_letter(col)}{row}"]
            if col == 7:  # Для столбца "Details" особое форматирование
                cell.alignment = Alignment(horizontal='justify', vertical='center', wrap_text=True)  # Выравнивание по ширине (justify)
            else:
                cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)  # Включаем перенос текста

        # Устанавливаем оптимальные размеры столбцов для текста
        ws.column_dimensions['B'].width = 40  # Столбец URL (шире для гиперссылок)
        ws.column_dimensions['D'].width = 25  # Столбец Title
        ws.column_dimensions['E'].width = 25  # Столбец Subtitle
        ws.column_dimensions['F'].width = 15  # Столбец Price
        ws.column_dimensions['G'].width = 60  # Столбец Details (с переносом текста)

    wb.save("messika_jewelry_data_with_images.xlsx")
    print("Данные успешно собраны и сохранены в messika_jewelry_data_with_images.xlsx")

# Основная функция
def collect_jewelry_data():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    wait = WebDriverWait(driver, 20)

    categories = [
        {"url": 'https://www.messika.com/en/jewelry/categories/diamond-ring', "name": "rings"},
        {"url": 'https://www.messika.com/en/jewelry/categories/diamond-bracelet', "name": "bracelets"},
        {"url": 'https://www.messika.com/en/jewelry/categories/diamond-earrings', "name": "earrings"},
        {"url": 'https://www.messika.com/en/jewelry/categories/diamond-necklace', "name": "necklaces"},
        {"url": 'https://www.messika.com/en/jewelry/categories/luxury-gifts', "name": "gifts"}
    ]

    all_data = []

    try:
        for category in categories:
            base_url = category["url"]
            category_name = category["name"]
            
            # Сбор данных для категории
            data = collect_data_for_category(driver, wait, base_url, category_name)
            all_data.extend(data)
    
    finally:
        driver.quit()

    # Сохранение данных в Excel
    save_data_to_excel(all_data)

if __name__ == "__main__":
    collect_jewelry_data()
